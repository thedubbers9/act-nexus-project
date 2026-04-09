#!/usr/bin/env python3
"""Generate minimal TAIDL + XLA → Gemmini mapping sheets and JSON interface."""

import csv
import json
from pathlib import Path

ISA_DIR = Path(__file__).resolve().parent
XLA_CLASS_CSV = ISA_DIR / "xla_taidl_hardware_class_view.csv"
TAIDL_JOIN_CSV = ISA_DIR / "taidl_xla_join_reference.csv"
OUT_XLSX = ISA_DIR / "gemmini_mapping_simple.xlsx"
OUT_JSON = ISA_DIR / "gemmini_mapping_interface.json"
OUT_TAIDL_CSV = ISA_DIR / "gemmini_mapping_simple_taidl.csv"
OUT_XLA_CSV = ISA_DIR / "gemmini_mapping_simple_xla.csv"

# hardware_abstraction_class -> default Gemmini mapping (minimal columns)
CLASS_DEFAULTS = {
    "tensor_compute": {
        "map_group": "tensor_mac",
        "gemmini_ips": "dma_load; dma_store; systolic_mesh; transposer",
        "multi_ip": "yes",
        "how_realized": "mvin/mvout + config_ex + matmul.preload/matmul.compute (or gemmini_loop_ws / gemmini_loop_conv_ws).",
        "act_rule": "After TAIDL→XLA normalize; bind tensors to SPAD/ACC/HBM; map Dot/Conv to execute path.",
    },
    "vector_compute": {
        "map_group": "vector_elemwise",
        "gemmini_ips": "accumulator_alu; rocket_core",
        "multi_ip": "sometimes",
        "how_realized": "Bias/residual often fused into matmul D/acc; else library loops on SPAD or Rocket.",
        "act_rule": "Prefer fusion with tensor ops in ACT lowering; else charge peripheral or host fallback.",
    },
    "reduction": {
        "map_group": "reduce",
        "gemmini_ips": "systolic_mesh; acc_sram; rocket_core",
        "multi_ip": "yes",
        "how_realized": "Sum-like: partials in ACC (WS) or passes; finalize may need Rocket.",
        "act_rule": "Decompose Reduce to supported combiners; map sum-like to acc path.",
    },
    "special_math": {
        "map_group": "sfu",
        "gemmini_ips": "config_norm; rocket_core",
        "multi_ip": "sometimes",
        "how_realized": "Subset via experimental config_norm (I-BERT style); general Exp/Log/etc → Rocket.",
        "act_rule": "If TAIDL expands to TAIDL primitives, map each; else host until SFU modeled.",
    },
    "contiguous_move": {
        "map_group": "dma_move",
        "gemmini_ips": "dma_load; dma_store",
        "multi_ip": "yes",
        "how_realized": "mvin / mvout with config_mvin / config_mvout strides.",
        "act_rule": "Layout lowering must expose byte/row strides for DMA cost.",
    },
    "logical_view": {
        "map_group": "layout_meta",
        "gemmini_ips": "controller_agu",
        "multi_ip": "no",
        "how_realized": "No RoCC op; metadata/addressing unless materialized → then DMA.",
        "act_rule": "Zero cost until ACT/backend forces copy (reshape/broadcast materialized).",
    },
    "predication_select": {
        "map_group": "compare_select",
        "gemmini_ips": "relu_config; mvout_maxpool; rocket_core",
        "multi_ip": "sometimes",
        "how_realized": "ReLU via config_ex; pool via mvout path; general select/compare → Rocket.",
        "act_rule": "Match fused activation/pool patterns before generic predication.",
    },
    "input_literal": {
        "map_group": "inputs",
        "gemmini_ips": "host_mem; dma_load",
        "multi_ip": "yes",
        "how_realized": "Constants/params in DRAM → mvin into SPAD/ACC.",
        "act_rule": "Phase: bind inputs; size tensors before movement cost.",
    },
    "indexed_move": {
        "map_group": "irregular",
        "gemmini_ips": "rocket_core",
        "multi_ip": "no",
        "how_realized": "No gather/scatter engine; software on host or scalar loops.",
        "act_rule": "Flag unsupported on-device path; do not charge systolic for indexed ops.",
    },
    "control_runtime": {
        "map_group": "control",
        "gemmini_ips": "rocket_core",
        "multi_ip": "no",
        "how_realized": "Host issues RoCC stream; tokens/async are compiler metadata.",
        "act_rule": "No Gemmini datapath cost; sequencing only.",
    },
    "collective_comm": {
        "map_group": "multi_chip",
        "gemmini_ips": "not_on_tile",
        "multi_ip": "no",
        "how_realized": "Not modeled for single Gemmini; map to external NOC/host.",
        "act_rule": "Out of scope for single-accelerator ACT profile unless extended.",
    },
    "io_transfer": {
        "map_group": "io",
        "gemmini_ips": "host_interface; dma",
        "multi_ip": "yes",
        "how_realized": "Treat as host↔device transfer analogous to mvin/mvout.",
        "act_rule": "Attach to IO bandwidth bucket, not systolic.",
    },
    "random_generation": {
        "map_group": "rng",
        "gemmini_ips": "rocket_core",
        "multi_ip": "no",
        "how_realized": "Host/scalar RNG; not Gemmini ISA.",
        "act_rule": "Host fallback.",
    },
    "specialized_compute": {
        "map_group": "domain_kernels",
        "gemmini_ips": "decompose; systolic_mesh; rocket_core",
        "multi_ip": "yes",
        "how_realized": "Lower to Dot/Conv/vector ops per ACT recipe; else Rocket.",
        "act_rule": "Expand to TAIDL primitives before hardware map.",
    },
    "tuple_structural": {
        "map_group": "ir_only",
        "gemmini_ips": "none",
        "multi_ip": "no",
        "how_realized": "Tuple packing; no tensor execution.",
        "act_rule": "Ignore for datapath cost.",
    },
}

XLA_OP_NOTES = {
    "Dot": "Primary systolic path: A*B+D→C.",
    "Conv (Convolution)": "Use gemmini_loop_conv_ws when available.",
    "Reduce": "Combiner-dependent; add on acc vs host.",
    "Select": "TAIDL helpers select_lt/select_eq_var decompose here.",
}


def row_for_xla_op(xla_op, hw_class, in_taidl, taidl_primitives):
    d = dict(CLASS_DEFAULTS.get(hw_class, CLASS_DEFAULTS["specialized_compute"]))
    note = XLA_OP_NOTES.get(xla_op)
    how = d["how_realized"]
    if note:
        how = f"{note} {how}"
    return {
        "xla_hlo_op": xla_op,
        "hw_abstraction_class": hw_class,
        "in_taidl": in_taidl,
        "taidl_primitives": taidl_primitives,
        "map_group": d["map_group"],
        "gemmini_ips": d["gemmini_ips"],
        "uses_multiple_ip": d["multi_ip"],
        "how_realized": how.strip(),
        "act_taidl_stage_rule": d["act_rule"],
    }


def load_taidl_join():
    with open(TAIDL_JOIN_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_xla_class_view():
    with open(XLA_CLASS_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# Per-primitive fixes (join-table group is sometimes too coarse for Gemmini).
TAIDL_PRIM_OVERRIDES = {
    "convert": {
        "map_group": "type_convert",
        "gemmini_ips": "mvin_scale; acc_scale_read; rocket_core",
        "uses_multiple_ip": "sometimes",
        "how_realized": "If quant: scale on mvin / acc read-down; else true cast → Rocket or explicit elementwise pass.",
        "act_taidl_stage_rule": "Split metadata-only vs value-changing convert before costing.",
    },
    "transpose": {
        "map_group": "layout_move",
        "gemmini_ips": "transposer; dma_load; dma_store",
        "uses_multiple_ip": "sometimes",
        "how_realized": "Prefer fuse into matmul (config_ex + transposer); else explicit data movement.",
        "act_taidl_stage_rule": "Do not add standalone transpose cost if folded into execute config.",
    },
}


def build_taidl_rows():
    rows = []
    key_map = {
        "layout_metadata": "logical_view",
        "elementwise_alu": "vector_compute",
        "elementwise_special": "special_math",
        "data_movement": "contiguous_move",
        "data_movement_update": "indexed_move",
        "tensor_compute": "tensor_compute",
        "reduction": "reduction",
        "predication_select": "predication_select",
        "input_literal": "input_literal",
    }
    for r in load_taidl_join():
        xla = r["matched_xla_hlo_op"]
        sg = r["suggested_abstraction_group"]
        base = CLASS_DEFAULTS.get(
            key_map.get(sg, "specialized_compute"),
            CLASS_DEFAULTS["specialized_compute"],
        )
        prim = r["taidl_primitive"]
        row = {
            "taidl_primitive": prim,
            "canonical_xla": xla,
            "map_group": base["map_group"],
            "gemmini_ips": base["gemmini_ips"],
            "uses_multiple_ip": base["multi_ip"],
            "how_realized": base["how_realized"],
            "act_taidl_stage_rule": base["act_rule"],
        }
        if prim in TAIDL_PRIM_OVERRIDES:
            row.update(TAIDL_PRIM_OVERRIDES[prim])
        rows.append(row)
    return rows


# Gemmini-specific fixes where the reference XLA class is misleading.
XLA_HW_CLASS_OVERRIDES = {
    "DynamicUpdateSlice": "indexed_move",
}


def build_xla_rows():
    out = []
    for r in load_xla_class_view():
        op = r["xla_hlo_op"]
        hw_class = XLA_HW_CLASS_OVERRIDES.get(op, r["hardware_abstraction_class"])
        out.append(
            row_for_xla_op(
                op,
                hw_class,
                str(r.get("in_taidl", "")),
                str(r.get("taidl_primitives", "")),
            )
        )
    return out


def write_csv(path, fieldnames, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(taidl_rows, xla_rows):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed; skipping .xlsx (CSVs written).")
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Rules"
    ws0.append(["rule_name", "text"])
    for name, text in [
        (
            "normalize",
            "Map raw TAIDL/HLO spellings to canonical primitives before Gemmini mapping.",
        ),
        (
            "fusion_first",
            "If a fused_pattern matches, use its IP list once; do not also charge each primitive.",
        ),
        (
            "multi_ip",
            "When uses_multiple_ip=yes, cost = sum of IP buckets (or shared schedule per your model).",
        ),
        (
            "layout_free",
            "layout_meta group: zero bytes until ACT forces materialize copy.",
        ),
        (
            "host_fallback",
            "rocket_core when Gemmini has no ISA for the op.",
        ),
    ]:
        ws0.append([name, text])

    ws1 = wb.create_sheet("TAIDL")
    h1 = list(taidl_rows[0].keys()) if taidl_rows else []
    ws1.append(h1)
    for row in taidl_rows:
        ws1.append([row[k] for k in h1])

    ws2 = wb.create_sheet("XLA_All")
    h2 = list(xla_rows[0].keys()) if xla_rows else []
    ws2.append(h2)
    for row in xla_rows:
        ws2.append([row[k] for k in h2])

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


def build_interface_json():
    return {
        "version": 1,
        "accelerator": "gemmini",
        "reference": "https://github.com/ucb-bar/gemmini",
        "ip_groups": {
            "tensor_mac": {
                "members": ["dma_load", "dma_store", "systolic_mesh", "transposer"],
                "note": "Typical Dot/Conv uses all four in sequence/overlap.",
            },
            "vector_elemwise": {
                "members": ["accumulator_alu", "rocket_core"],
                "note": "Often fused into matmul epilogue; else peripheral or host.",
            },
            "layout_meta": {
                "members": ["controller_agu"],
                "note": "Metadata-only unless materialized.",
            },
            "dma_move": {
                "members": ["dma_load", "dma_store"],
                "note": "mvin / mvout family.",
            },
        },
        "act_pipeline_stages": [
            {
                "stage": "1_taidl_semantics",
                "role": "Instruction means TAIDL/XLA-semantic ops (no hardware yet).",
            },
            {
                "stage": "2_normalize",
                "role": "Expand aliases/helpers (e.g. exp, select_*) to canonical primitive graph.",
            },
            {
                "stage": "3_bind_storage",
                "role": "Assign SPAD/ACC/DRAM per ACT instruction binding (drives DMA bytes).",
            },
            {
                "stage": "4_map_to_gemmini",
                "role": "Apply fused_patterns then primitive_to_ip; emit multi-IP rows when needed.",
            },
        ],
        "precedence_rules": [
            "fused_patterns checked in list order before primitive_to_ip fallback",
            "Do not double-count transpose folded into matmul config",
            "collective_comm and tuple_structural excluded from systolic cost by default",
        ],
        "fused_patterns": [
            {
                "name": "matmul_bias_relu",
                "match_primitives": ["dot", "add", "maximum"],
                "gemmini_ips": ["dma_load", "dma_store", "systolic_mesh", "relu_config"],
                "note": "Bias as D in OS matmul; ReLU via config_ex when fused by lowering.",
            },
            {
                "name": "conv_ws_loop",
                "match_xla": ["Conv (Convolution)"],
                "gemmini_ips": ["dma_load", "dma_store", "systolic_mesh", "loop_conv_fsm"],
                "note": "Single CISC region gemmini_loop_conv_ws when compiler emits it.",
            },
        ],
        "primitive_to_ip_fallback": {
            "dot": ["dma_load", "dma_store", "systolic_mesh", "transposer"],
            "add": ["accumulator_alu", "rocket_core"],
            "reduce_add": ["systolic_mesh", "acc_sram", "rocket_core"],
            "broadcast": ["controller_agu"],
            "copy": ["dma_load", "dma_store"],
        },
    }


def main():
    taidl_rows = build_taidl_rows()
    xla_rows = build_xla_rows()

    if taidl_rows:
        write_csv(OUT_TAIDL_CSV, list(taidl_rows[0].keys()), taidl_rows)
        print(f"Wrote {OUT_TAIDL_CSV}")
    write_csv(OUT_XLA_CSV, list(xla_rows[0].keys()), xla_rows)
    print(f"Wrote {OUT_XLA_CSV}")

    write_xlsx(taidl_rows, xla_rows)

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(build_interface_json(), f, indent=2)
        f.write("\n")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
