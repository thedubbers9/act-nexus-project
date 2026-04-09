#!/usr/bin/env python3
"""Generate the final realization-first hardware-mapping artefacts.

Reads the three source CSVs and the fused-patterns JSON, then emits:
  - final_mapping.xlsx   (multi-sheet workbook)
  - final_mapping.json   (machine-readable interface)

Requires: openpyxl  (pip install --user openpyxl)
"""

import csv
import json
import os
import sys
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("openpyxl is required: python3 -m pip install --user openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(HERE, "final_mapping.xlsx")
OUT_JSON = os.path.join(HERE, "final_mapping.json")

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DEFAULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def read_csv(name):
    path = os.path.join(HERE, name)
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def read_json(name):
    path = os.path.join(HERE, name)
    with open(path) as f:
        return json.load(f)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws, nrows, ncols, highlight_col=None):
    for r in range(2, nrows + 1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        if highlight_col is not None:
            val = ws.cell(row=r, column=highlight_col).value
            if val and str(val).strip() == "1":
                fill = DEFAULT_FILL
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def auto_width(ws, ncols):
    for col in range(1, ncols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)


def write_sheet(ws, rows, columns, highlight_col_name=None):
    for ci, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col_name)
    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(columns, 1):
            ws.cell(row=ri, column=ci, value=row.get(col_name, ""))
    ncols = len(columns)
    nrows = len(rows) + 1
    highlight_col = None
    if highlight_col_name:
        highlight_col = columns.index(highlight_col_name) + 1
    style_header(ws, ncols)
    style_body(ws, nrows, ncols, highlight_col)
    auto_width(ws, ncols)


def build_ip_blocks_for_realization(ip_flow_rows, rid):
    """Return ordered list of ip_block names for a realization_id."""
    steps = sorted(
        [r for r in ip_flow_rows if r["realization_id"] == rid],
        key=lambda r: int(r["step"]),
    )
    return [s["ip_block"] for s in steps]


def build_ip_flow_for_realization(ip_flow_rows, rid):
    steps = sorted(
        [r for r in ip_flow_rows if r["realization_id"] == rid],
        key=lambda r: int(r["step"]),
    )
    return [
        {"step": int(s["step"]), "ip_block": s["ip_block"],
         "direction": s["direction"], "notes": s.get("notes", "")}
        for s in steps
    ]


def main():
    prim_rows = read_csv("primitive_realizations.csv")
    ip_flow_rows = read_csv("realization_ip_flow.csv")
    cost_rows = read_csv("realization_cost_tags.csv")
    fused_data = read_json("fused_patterns.json")

    cost_map = {}
    for r in cost_rows:
        cost_map[r["realization_id"]] = {
            "cost_tag": r["cost_tag"],
            "cost_formula_hint": r.get("cost_formula_hint", ""),
            "notes": r.get("notes", ""),
        }

    # ---- augment prim_rows with ip_blocks and cost_tag for the spreadsheet ----
    for row in prim_rows:
        rid = row["realization_id"]
        blocks = build_ip_blocks_for_realization(ip_flow_rows, rid)
        row["ip_blocks"] = " → ".join(blocks)
        cm = cost_map.get(rid, {})
        row["cost_tag"] = cm.get("cost_tag", "")
        row["cost_formula_hint"] = cm.get("cost_formula_hint", "")

    # ---- build xlsx ----
    wb = Workbook()

    # Sheet 1: README / Rules
    ws_rules = wb.active
    ws_rules.title = "README"
    rules_text = [
        ["Realization-First Hardware Mapping Interface"],
        [""],
        ["Schema: primitive → realization → IP flow → cost tag"],
        [""],
        ["Precedence rules:"],
        ["1. Fused patterns (fused_patterns.json) are checked first, highest priority wins."],
        ["2. Unmatched primitives use per-op fallback from primitive_realizations."],
        ["3. For each primitive, the row marked is_default=1 is used unless a condition selects an alternative."],
        ["4. Each primitive instance is assigned to exactly one realization."],
        [""],
        ["Sheets:"],
        ["  Realizations — each TAIDL primitive and its possible realizations"],
        ["  IP_Flow      — ordered IP pipeline for each realization"],
        ["  Cost_Tags    — accounting tag and formula hint for each realization"],
        ["  Fused        — fusion overrides with priority"],
        [""],
        ["Machine-readable: final_mapping.json"],
    ]
    for ri, line in enumerate(rules_text, 1):
        ws_rules.cell(row=ri, column=1, value=line[0])
        ws_rules.cell(row=ri, column=1).font = Font(name="Calibri", size=11)
        if ri == 1:
            ws_rules.cell(row=ri, column=1).font = Font(name="Calibri", size=14, bold=True)
    ws_rules.column_dimensions["A"].width = 90

    # Sheet 2: Realizations
    ws_real = wb.create_sheet("Realizations")
    real_cols = [
        "primitive", "realization_id", "is_default", "condition",
        "ip_blocks", "cost_tag", "cost_formula_hint", "gemmini_rationale",
    ]
    write_sheet(ws_real, prim_rows, real_cols, highlight_col_name="is_default")

    # Sheet 3: IP Flow
    ws_flow = wb.create_sheet("IP_Flow")
    flow_cols = ["realization_id", "step", "ip_block", "direction", "notes"]
    write_sheet(ws_flow, ip_flow_rows, flow_cols)

    # Sheet 4: Cost Tags
    ws_cost = wb.create_sheet("Cost_Tags")
    cost_cols = ["realization_id", "cost_tag", "cost_formula_hint", "notes"]
    write_sheet(ws_cost, cost_rows, cost_cols)

    # Sheet 5: Fused patterns
    ws_fused = wb.create_sheet("Fused")
    fused_flat = []
    for pat in fused_data["fused_patterns"]:
        fused_flat.append({
            "name": pat["name"],
            "match": " + ".join(pat["match"]),
            "realization_id": pat["realization_id"],
            "priority": str(pat["priority"]),
            "cost_tag": pat["cost_tag"],
            "ip_flow_summary": " → ".join(s["ip_block"] for s in pat["ip_flow"]),
            "notes": pat.get("notes", ""),
        })
    fused_cols = ["name", "match", "realization_id", "priority",
                  "cost_tag", "ip_flow_summary", "notes"]
    write_sheet(ws_fused, fused_flat, fused_cols)

    wb.save(OUT_XLSX)
    print("Wrote", OUT_XLSX)

    # ---- build json ----
    primitives_json = OrderedDict()
    for row in prim_rows:
        prim = row["primitive"]
        if prim not in primitives_json:
            primitives_json[prim] = {"realizations": []}
        rid = row["realization_id"]
        entry = {
            "realization_id": rid,
            "is_default": row["is_default"] == "1",
            "condition": row["condition"],
            "ip_flow": build_ip_flow_for_realization(ip_flow_rows, rid),
            "cost_tag": cost_map.get(rid, {}).get("cost_tag", ""),
            "cost_formula_hint": cost_map.get(rid, {}).get("cost_formula_hint", ""),
            "gemmini_rationale": row.get("gemmini_rationale", ""),
        }
        primitives_json[prim]["realizations"].append(entry)

    out = OrderedDict()
    out["version"] = 3
    out["model"] = "realization_first"
    out["description"] = (
        "Each TAIDL/XLA primitive maps to one or more realizations. "
        "Each realization maps to an ordered IP flow and a cost tag. "
        "Fused patterns override individual per-op mappings."
    )
    out["pipeline_stages"] = [
        "1. Semantics: TAIDL/XLA meaning",
        "2. Primitive decomposition",
        "3. Normalization (canonical op names)",
        "4. Fusion check (fused_patterns, highest priority first)",
        "5. Per-op realization selection (default unless condition selects alternative)",
        "6. IP flow assignment",
        "7. Cost aggregation via cost_tag",
    ]
    out["primitives"] = primitives_json
    out["fused_patterns"] = fused_data["fused_patterns"]

    ip_catalog = OrderedDict()
    seen_ips = set()
    for row in ip_flow_rows:
        ip = row["ip_block"]
        if ip not in seen_ips and ip != "none":
            seen_ips.add(ip)
            ip_catalog[ip] = {"gemmini_block": ip, "notes": row.get("notes", "")}
    out["ip_catalog"] = ip_catalog

    cost_tag_catalog = OrderedDict()
    for row in cost_rows:
        tag = row["cost_tag"]
        if tag not in cost_tag_catalog:
            cost_tag_catalog[tag] = {
                "description": row.get("notes", ""),
                "formula_hint": row.get("cost_formula_hint", ""),
            }
    out["cost_tag_catalog"] = cost_tag_catalog

    with open(OUT_JSON, "w") as f:
        json.dump(out, f, indent=2)
    print("Wrote", OUT_JSON)

    # ---- summary ----
    n_prims = len(set(r["primitive"] for r in prim_rows))
    n_realizations = len(set(r["realization_id"] for r in prim_rows))
    n_ips = len(seen_ips)
    n_fused = len(fused_data["fused_patterns"])
    print("\nSummary:")
    print("  Primitives:    ", n_prims)
    print("  Realizations:  ", n_realizations)
    print("  IP blocks:     ", n_ips)
    print("  Fused patterns:", n_fused)


if __name__ == "__main__":
    main()
